import json
import requests

from shuffle_sdk import AppBase
from shuffle_sdk import csv_parse

from openpyxl import Workbook, load_workbook

class MSExcel(AppBase):
    __version__ = "1.0.0"
    app_name = "Microsoft Excel"

    def __init__(self, redis, logger, console_logger=None):
        """
        Each app should have this __init__ to set up Redis and logging.
        :param redis:
        :param logger:
        :param console_logger:
        """
        super().__init__(redis, logger, console_logger)

    def authenticate(self, tenant_id, client_id, client_secret, graph_url):
        s = requests.Session()
        auth_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        auth_data = {
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": f"{graph_url}/.default",
        }
        auth_headers = {
            "Content-Type": "application/x-www-form-urlencoded",
            "cache-control": "no-cache",
        }

        print(f"Making request to: {auth_url}")
        res = s.post(auth_url, data=auth_data, headers=auth_headers)

        # Auth failed, raise exception with the response
        if res.status_code != 200:
            raise ConnectionError(res.text)

        access_token = res.json().get("access_token")
        s.headers = {"Authorization": f"Bearer {access_token}", "cache-control": "no-cache"}
        return s

    def get_user_id(self, tenant_id, client_id, client_secret):
        graph_url = "https://graph.microsoft.com"
        session = self.authenticate(tenant_id, client_id, client_secret, graph_url)
        graph_url = "https://graph.microsoft.com/v1.0/users"
        ret = session.get(graph_url)
        return ret.text

    def get_files(self, tenant_id, client_id, client_secret, user_id):
        graph_url = "https://graph.microsoft.com"
        session = self.authenticate(tenant_id, client_id, client_secret, graph_url)
        graph_url = f"https://graph.microsoft.com/v1.0/users/{user_id}/drive/root/children"
        ret = session.get(graph_url)
        return ret.text

    def list_worksheets(self, tenant_id, client_id, client_secret, user_id, file_id):
        graph_url = "https://graph.microsoft.com"
        session = self.authenticate(tenant_id, client_id, client_secret, graph_url)
        graph_url = f"https://graph.microsoft.com/v1.0/users/{user_id}/drive/items/{file_id}/workbook/worksheets"
        ret = session.get(graph_url)
        return ret.text

    def add_worksheet(self, tenant_id, client_id, client_secret, user_id, file_id, name):
        graph_url = "https://graph.microsoft.com"
        session = self.authenticate(tenant_id, client_id, client_secret, graph_url)
        graph_url = f"https://graph.microsoft.com/v1.0/users/{user_id}/drive/items/{file_id}/workbook/worksheets"
        if len(name)!=0:
            body = {
                "name": name
                }
        else:
            body = {}
        ret = session.post(graph_url, json = body)
        return ret.text

    def delete_worksheet(self, tenant_id, client_id, client_secret, user_id, file_id, name):
        graph_url = "https://graph.microsoft.com"
        session = self.authenticate(tenant_id, client_id, client_secret, graph_url)
        graph_url = f"https://graph.microsoft.com/v1.0/users/{user_id}/drive/items/{file_id}/workbook/worksheets/{name}"
        ret = session.delete(graph_url)
        if ret.status_code != 200:
            return "Action failed"
        else:
            return "Action successfully completed"
        
    def insert_or_update_data(self, tenant_id, client_id, client_secret, user_id, file_id, sheet_name, address, value):
        graph_url = "https://graph.microsoft.com"
        session = self.authenticate(tenant_id, client_id, client_secret, graph_url)
        graph_url = f"https://graph.microsoft.com/v1.0/users/{user_id}/drive/items/{file_id}/workbook/worksheets/{sheet_name}/range(address='{address}')"
        lt = []
        for i in value.split(';'):
            temp_var = []
            for j in i.split(','):
                temp_var.append(j)
            lt.append(temp_var)
        body = {
            "values":lt
        }
        ret = session.patch(graph_url, json=body)
        return ret.text

    def clear_data(self, tenant_id, client_id, client_secret, user_id, file_id, sheet_name, address):
        graph_url = "https://graph.microsoft.com"
        session = self.authenticate(tenant_id, client_id, client_secret, graph_url)
        graph_url = f"https://graph.microsoft.com/v1.0/users/{user_id}/drive/items/{file_id}/workbook/worksheets/{sheet_name}/range(address='{address}')/clear"
        ret = session.post(graph_url)
        if ret.status_code != 200:
            return "Action failed"
        else:
            return "Action successfully completed"

    def convert_to_csv(self, tenant_id, client_id, client_secret, file_id, sheet="Sheet1"):
        filedata = self.get_file(file_id)
        if filedata["success"] != True:
            return filedata
    
        basename = "/tmp/file.xlsx"
        with open(basename, "wb") as tmp:
            tmp.write(filedata["data"])
    
        if sheet == "":
            sheet = "Sheet1"
    
        #wb = Workbook(basename)
        wb = load_workbook(basename, read_only=True)

        # grab the active worksheet
        ws = wb.active
        #for item in ws.iter_rows():
        #    print(item)
    
        csvdata = ""
        for row in ws.values:
            for value in row:
                #print(value)
                if value == None:
                    csvdata += ","
                elif isinstance(value, str):
                    csvdata += value+","
                else:
                    csvdata += str(value)+","
    
            csvdata = csvdata[:-1]+"\n"
        csvdata = csvdata[:-1]
    
        print("Data length: (%s)" % len(csvdata))

        return csvdata

    def get_excel_file_data(self, file_id, to_list=True, sheets="", max_rows=100000, skip_rows=0):
        filedata = self.get_file(file_id)
        if filedata["success"] != True:
            print(f"[ERROR] Bad info from file: {filedata}") 
            return filedata

        if not sheets:
            sheets = ""

        sheets = sheets.lower()
        max_rows = int(max_rows)
        skip_rows = int(skip_rows)

        try:
            #print("Filename: %s" % filedata["filename"])
            if "csv" in filedata["filename"]:
                try:
                    filedata["data"] = filedata["data"].decode("utf-8")
                except:
                    try:
                        filedata["data"] = filedata["data"].decode("utf-16")
                    except:
                        filedata["data"] = filedata["data"].decode("latin-1")

                returndata = csv_parse(filedata["data"])
                return returndata

        except Exception as e:
            print("Error parsing file with csv parser for file %s: %s" % (filedata["filename"], e))
    
        basename = "/tmp/file.xlsx"
        with open(basename, "wb") as tmp:
            tmp.write(filedata["data"])
    
        #wb = Workbook(basename)
        try:
            wb = load_workbook(basename, read_only=True)
        except Exception as e:
            return {
                "success": False,
                "reason": "The file is invalid. Are you sure it's a valid excel file? CSV files are not supported.",
                "exception": "Error: %s" % e,
            }
    
        # Default
        #max_count = 25000
        #if os.getenv("SHUFFLE_APP_SDK_TIMEOUT") > 240:
        # Limits are ~no longer relevant if to_list=True

        cnt = 0
        skipped_cnt = 0 
        output_data = []
        for ws in wb.worksheets:
            if ws.title.lower() not in sheets and sheets != "":
                continue
    
            # grab the active worksheet
            csvdata = ""
            if cnt-skipped_cnt > skip_rows:
                break

            list_data = []
            for row in ws.values:
                cnt += 1
                if cnt < skip_rows:
                    skipped_cnt += 1
                    continue

                if cnt-skipped_cnt > max_rows:
                    break

                for value in row:
                    #print(value)
                    if value == None:
                        csvdata += ","
                    elif isinstance(value, str):
                        csvdata += value+","
                    else:
                        csvdata += str(value)+","
    
                list_data.append(csvdata)
                if to_list == False:
                    csvdata = csvdata[:-1]+"\n"
                else:
                    csvdata = ""

            #csvdata = csvdata[:-1]

            output = {
                "sheet": ws.title,
                "data": csvdata,
            }

            if to_list == False:
                print("Data len (%s): %d" % (ws.title, len(csvdata)))
                output_data.append(output)
            else:
                print("Data len (%s): %d" % (ws.title, len(list_data)))
                output_data.append({
                    "sheet": ws.title,
                    "data": list_data,
                })

        print("Done! Returning data of length: %d" % len(output_data))
    
        return output_data
        
if __name__ == "__main__":
    MSExcel.run()
